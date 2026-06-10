import os
from datetime import date, datetime, timedelta
from collections import defaultdict

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.models.employee import Employee
from app.models.vacation import Vacation
from app.models.calendar import CalendarDay, YearCalendar
from app.services.calendar_service import get_calendar_days, get_transfers_for_year
from app.utils.auth import require_auth_page

router = APIRouter(tags=["pages"])
templates = Jinja2Templates(directory="app/templates")


def _detect_overlaps(vacations: list[dict], year: int) -> tuple[dict, set[int], list[dict]]:
    """
    Анализирует пересечения отпусков.
    Возвращает:
      - day_load: {день_года: кол-во_сотрудников_в_отпуске}
      - overlapping_vacation_ids: set(id отпусков, которые пересекаются)
      - conflicts: список конфликтов [{days, employees, start, end}]
    """
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    total_days = (year_end - year_start).days + 1

    # day_load: для каждого дня считаем сколько сотрудников в отпуске
    day_load = defaultdict(lambda: {"count": 0, "employee_ids": set(), "vacation_ids": set()})

    for v in vacations:
        s = date.fromisoformat(v["start_date"])
        e = date.fromisoformat(v["end_date"])
        display_s = max(s, year_start)
        display_e = min(e, year_end)

        current = display_s
        while current <= display_e:
            day_key = (current - year_start).days
            day_load[day_key]["count"] += 1
            day_load[day_key]["employee_ids"].add(v["employee_id"])
            day_load[day_key]["vacation_ids"].add(v["id"])
            current += timedelta(days=1)

    # Дни с пересечением (≥2 сотрудников)
    overlap_days = {d for d, info in day_load.items() if info["count"] >= 2}

    # Какие отпуска затронуты пересечениями
    overlapping_vacation_ids = set()
    for d in overlap_days:
        overlapping_vacation_ids.update(day_load[d]["vacation_ids"])

    # Группируем смежные дни конфликтов в периоды
    conflicts = []
    if overlap_days:
        sorted_days = sorted(overlap_days)
        period_start = sorted_days[0]
        period_employees = day_load[period_start]["employee_ids"].copy()

        for i in range(1, len(sorted_days)):
            day = sorted_days[i]
            if day == sorted_days[i - 1] + 1:
                period_employees.update(day_load[day]["employee_ids"])
            else:
                # Конец периода
                conflicts.append({
                    "start_day": period_start,
                    "end_day": sorted_days[i - 1],
                    "start_date": str(year_start + timedelta(days=period_start)),
                    "end_date": str(year_start + timedelta(days=sorted_days[i - 1])),
                    "employees_count": len(period_employees),
                })
                period_start = day
                period_employees = day_load[day]["employee_ids"].copy()

        # Последний период
        conflicts.append({
            "start_day": period_start,
            "end_day": sorted_days[-1],
            "start_date": str(year_start + timedelta(days=period_start)),
            "end_date": str(year_start + timedelta(days=sorted_days[-1])),
            "employees_count": len(period_employees),
        })

    # Преобразуем day_load в словарь {день: count} для шаблона
    day_load_simple = {d: info["count"] for d, info in day_load.items()}

    return day_load_simple, overlapping_vacation_ids, conflicts

# ── Месяцы для отображения ────────────────────────────────────────────────

MONTH_NAMES = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]


# ── Страницы ──────────────────────────────────────────────────────────────

@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Страница входа."""
    return templates.TemplateResponse("login.html", {"request": request})


@router.get("/", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    year: int = Query(default_factory=lambda: date.today().year),
    db: Session = Depends(get_db),
):
    """Главная страница — дашборд с диаграммой Ганта."""
    admin, redirect = require_auth_page(request, db)
    if redirect:
        return redirect

    employees = (
        db.query(Employee)
        .filter(Employee.is_active == True)
        .options(selectinload(Employee.vacations))
        .order_by(Employee.full_name)
        .all()
    )

    # Собираем все отпуска за выбранный год (для поиска пересечений)
    all_vacations_raw = []
    for emp in employees:
        for v in emp.vacations:
            if v.status == "cancelled":
                continue
            v_start_year = v.start_date.year
            v_end_year = v.end_date.year
            if v_end_year < year or v_start_year > year:
                continue
            all_vacations_raw.append({
                "id": v.id,
                "employee_id": emp.id,
                "employee_name": emp.full_name,
                "start_date": str(v.start_date),
                "end_date": str(v.end_date),
                "days_count": v.days_count,
                "status": v.status,
                "status_label": v.status_label,
            })

    # Детектируем пересечения
    day_load, overlapping_ids, conflicts = _detect_overlaps(all_vacations_raw, year)

    # Собираем отпуска за выбранный год (для отображения)
    vacations_data = []
    for emp in employees:
        for v in emp.vacations:
            if v.status == "cancelled":
                continue
            # Определяем, попадает ли отпуск в выбранный год
            v_start_year = v.start_date.year
            v_end_year = v.end_date.year

            if v_end_year < year or v_start_year > year:
                continue

            # Обрезаем отображение под год
            display_start = v.start_date
            display_end = v.end_date
            if v_start_year < year:
                display_start = date(year, 1, 1)
            if v_end_year > year:
                display_end = date(year, 12, 31)

            # Вычисляем процентное положение в году
            year_start = date(year, 1, 1)
            year_end = date(year, 12, 31)
            total_days = (year_end - year_start).days + 1

            left_pct = max(0, (display_start - year_start).days / total_days * 100)
            width_pct = max(1, (display_end - display_start).days / total_days * 100)

            vacations_data.append({
                "id": v.id,
                "employee_id": emp.id,
                "employee_name": emp.full_name,
                "start_date": str(v.start_date),
                "end_date": str(v.end_date),
                "days_count": v.days_count,
                "status": v.status,
                "status_label": v.status_label,
                "left_pct": round(left_pct, 2),
                "width_pct": round(width_pct, 2),
                "has_overlap": v.id in overlapping_ids,
            })

    # Группируем по сотрудникам
    gantt_data = defaultdict(list)
    for vd in vacations_data:
        gantt_data[vd["employee_id"]].append(vd)

    employee_vacations = [
        {
            "id": emp.id,
            "name": emp.full_name,
            "position": emp.position,
            "department": emp.department,
            "color": emp.color,
            "annual_days": emp.annual_days,
            "total_days": emp.total_days,
            "used_days": emp.used_days,
            "remaining_days": emp.remaining_days,
            "vacations": gantt_data.get(emp.id, []),
        }
        for emp in employees
    ]

    # Загружаем праздничные дни для календаря
    calendar_days = get_calendar_days(db, year)
    holiday_dates_cache = {d["day"] for d in calendar_days if d["is_holiday"]}

    # Готовим данные для сетки выходных (365 дней: для каждого дня — выходной, праздник или рабочий)
    weekend_grid = {}
    year_start = date(year, 1, 1)
    for day_offset in range(total_days):
        d = year_start + timedelta(days=day_offset)
        date_str = str(d)
        is_hol = date_str in holiday_dates_cache
        is_weekend = d.weekday() >= 5
        weekend_grid[day_offset] = is_hol or is_weekend
    # Строим сетку выходных/праздников по месяцам
    months_holidays = defaultdict(set)
    for d_str in holiday_dates_cache:
        try:
            d = datetime.strptime(d_str, "%Y-%m-%d").date()
            months_holidays[d.month].add(d.day)
        except ValueError:
            pass

    # Текущий месяц
    today = date.today()
    current_month = today.month if today.year == year else 1
    today_day_offset = (today - year_start).days if today.year == year else -1

    # Формируем данные для бокового календаря
    months_data = []
    for m in range(1, 13):
        month_holidays = months_holidays.get(m, set())
        months_data.append({
            "num": m,
            "name": MONTH_NAMES[m],
            "holiday_count": len(month_holidays),
            "is_current": m == current_month,
        })

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "employee_vacations": employee_vacations,
        "months_data": months_data,
        "current_year": year,
        "next_year": year + 1,
        "prev_year": year - 1,
        "today": str(today),
        "today_day_offset": today_day_offset,
        "month_names": MONTH_NAMES,
        "day_load": day_load,
        "total_days": 365,
        "conflicts": conflicts,
        "overlapping_count": len(conflicts),
        "weekend_grid": weekend_grid,
    })


@router.get("/export/excel")
async def export_excel(
    request: Request,
    year: int = Query(default_factory=lambda: date.today().year),
    db: Session = Depends(get_db),
):
    """Экспорт графика отпусков в Excel."""
    admin, redirect = require_auth_page(request, db)
    if redirect:
        return redirect

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    employees = (
        db.query(Employee)
        .filter(Employee.is_active == True)
        .options(selectinload(Employee.vacations))
        .order_by(Employee.full_name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"График отпусков {year}"

    # Заголовки
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Сотрудник", "Отдел", "Должность", "Лимит", "Перенос", "Всего", "Использовано", "Остаток", "Отпуска"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for row, emp in enumerate(employees, 2):
        vacations_str = "; ".join(
            f"{v.start_date}–{v.end_date} ({v.days_count} дн.)" for v in emp.vacations if v.status != "cancelled"
        ) or "—"

        ws.cell(row=row, column=1, value=emp.full_name)
        ws.cell(row=row, column=2, value=emp.department)
        ws.cell(row=row, column=3, value=emp.position)
        ws.cell(row=row, column=4, value=emp.annual_days)
        ws.cell(row=row, column=5, value=emp.carry_over_days)
        ws.cell(row=row, column=6, value=emp.total_days)
        ws.cell(row=row, column=7, value=emp.used_days)
        ws.cell(row=row, column=8, value=emp.remaining_days)
        ws.cell(row=row, column=9, value=vacations_str)

    # Ширина колонок
    widths = [30, 20, 25, 8, 8, 8, 12, 8, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"otpuska_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/employees", response_class=HTMLResponse)
async def employees_page(request: Request, db: Session = Depends(get_db)):
    """Страница списка сотрудников."""
    admin, redirect = require_auth_page(request, db)
    if redirect:
        return redirect

    employees = db.query(Employee).order_by(Employee.full_name).all()
    return templates.TemplateResponse("employees.html", {
        "request": request,
        "employees": [
            {
                "id": e.id,
                "full_name": e.full_name,
                "position": e.position,
                "department": e.department,
                "color": e.color,
                "annual_days": e.annual_days,
                "carry_over_days": e.carry_over_days,
                "total_days": e.total_days,
                "used_days": e.used_days,
                "remaining_days": e.remaining_days,
                "is_active": e.is_active,
                "vacation_count": len(e.vacations),
            }
            for e in employees
        ],
    })


@router.get("/employees/{employee_id}", response_class=HTMLResponse)
async def employee_detail(
    request: Request,
    employee_id: int,
    db: Session = Depends(get_db),
):
    """Страница деталей сотрудника с его отпусками."""
    admin, redirect = require_auth_page(request, db)
    if redirect:
        return redirect

    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        return RedirectResponse(url="/employees", status_code=302)

    vacations = (
        db.query(Vacation)
        .filter(Vacation.employee_id == employee_id)
        .order_by(Vacation.start_date.desc())
        .all()
    )

    total_d = employee.total_days
    used_d = employee.used_days
    prog_width = round((used_d / total_d * 100), 1) if total_d > 0 else 0

    return templates.TemplateResponse("employee_detail.html", {
        "request": request,
        "employee": {
            "id": employee.id,
            "full_name": employee.full_name,
            "position": employee.position,
            "department": employee.department,
            "color": employee.color,
            "annual_days": employee.annual_days,
            "total_days": total_d,
            "carry_over_days": employee.carry_over_days,
            "used_days": used_d,
            "remaining_days": employee.remaining_days,
            "prog_width": prog_width,
            "is_active": employee.is_active,
        },
        "vacations": [
            {
                "id": v.id,
                "start_date": str(v.start_date),
                "end_date": str(v.end_date),
                "days_count": v.days_count,
                "status": v.status,
                "status_label": v.status_label,
                "comment": v.comment,
            }
            for v in vacations
        ],
        "today": str(date.today()),
    })


@router.get("/calendar", response_class=HTMLResponse)
async def calendar_page(
    request: Request,
    year: int = Query(default_factory=lambda: date.today().year),
    db: Session = Depends(get_db),
):
    """Страница производственного календаря."""
    admin, redirect = require_auth_page(request, db)
    if redirect:
        return redirect

    calendar_days = get_calendar_days(db, year)
    has_data = len(calendar_days) > 0

    # Группируем по месяцам
    months = defaultdict(list)
    for d in calendar_days:
        try:
            dt = datetime.strptime(d["day"], "%Y-%m-%d").date()
            months[dt.month].append({
                "day": dt.day,
                "is_holiday": d["is_holiday"],
                "is_short_day": d["is_short_day"],
                "weekday": dt.weekday(),
                "holiday_name": d.get("holiday_name", ""),
            })
        except ValueError:
            continue

    # Строим календарную сетку по месяцам
    months_data = []
    for m in range(1, 13):
        days = months.get(m, [])
        work_days = sum(1 for d in days if not d["is_holiday"])
        holidays = sum(1 for d in days if d["is_holiday"])
        short = sum(1 for d in days if d["is_short_day"])

        months_data.append({
            "num": m,
            "name": MONTH_NAMES[m],
            "total_days": len(days),
            "work_days": work_days,
            "holidays": holidays,
            "short_days": short,
            "days": days,
        })

    transfers = get_transfers_for_year(year) if has_data else []

    return templates.TemplateResponse("calendar.html", {
        "request": request,
        "year": year,
        "next_year": year + 1,
        "prev_year": year - 1,
        "has_data": has_data,
        "months_data": months_data,
        "month_names": MONTH_NAMES,
        "transfers": transfers,
    })


# ── PWA ────────────────────────────────────────────────────────────────────

def _get_sw_js() -> str:
    cache_version = os.getenv("CACHE_VERSION", datetime.now().strftime("%Y%m%d%H%M%S"))
    cache_line = 'const CACHE = "otpysk-' + cache_version + '";\n'
    return cache_line + """
self.addEventListener("install", (e) => {
  self.skipWaiting();
});

self.addEventListener("activate", (e) => {
  e.waitUntil(
    caches.keys().then((keys) => Promise.all(
      keys.filter((k) => k !== CACHE).map((k) => caches.delete(k))
    ))
  );
  self.clients.claim();
});

self.addEventListener("fetch", (e) => {
  const url = new URL(e.request.url);
  const isNavigation = e.request.mode === "navigate";
  const isSameOrigin = url.origin === location.origin;
  const isStatic = url.pathname.match(/\\.(js|css|png|jpg|svg|ico|woff2?)$/i);

  if (isNavigation && isSameOrigin) {
    // HTML-страницы: всегда свежие (network-first с fallback на кэш)
    e.respondWith(
      fetch(e.request).then((response) => {
        const cloned = response.clone();
        caches.open(CACHE).then((cache) => cache.put(e.request, cloned));
        return response;
      }).catch(() => caches.match(e.request))
    );
  } else if (isStatic) {
    // Статика: cache-first
    e.respondWith(
      caches.match(e.request).then((cached) => {
        const fetched = fetch(e.request).then((response) => {
          if (response.ok) {
            const cloned = response.clone();
            caches.open(CACHE).then((cache) => cache.put(e.request, cloned));
          }
          return response;
        });
        return cached || fetched;
      })
    );
  }
  // Остальное (API, htmx CDN) — без кэширования
});
"""

_MANIFEST = """{
  "name": "Отпускной",
  "short_name": "Отпускной",
  "description": "Сервис учёта отпусков сотрудников",
  "start_url": "/",
  "display": "standalone",
  "background_color": "#f9fafb",
  "theme_color": "#1e293b",
  "orientation": "any",
  "lang": "ru",
  "icons": [
    {
      "src": "data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect width='100' height='100' rx='20' fill='%231e293b'/><text x='50%25' y='55%25' dominant-baseline='middle' text-anchor='middle' font-size='50'>🏖️</text></svg>",
      "sizes": "100x100",
      "type": "image/svg+xml",
      "purpose": "any maskable"
    }
  ]
}
"""


@router.get("/sw.js", response_class=PlainTextResponse)
async def service_worker():
    return PlainTextResponse(_get_sw_js(), media_type="application/javascript")


@router.get("/manifest.json")
async def manifest():
    import json
    return JSONResponse(json.loads(_MANIFEST), media_type="application/manifest+json")
